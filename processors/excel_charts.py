import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tempfile
import os

def generar_graficos_historial(historial_df, archivo_excel, log_func=None):
    if historial_df.empty:
        log_func and log_func("No hay datos para graficar.")
        return

    resumen_mes = historial_df.groupby(["Mes", "Tipo de Cambio"]).size().unstack(fill_value=0)

    plt.figure(figsize=(10, 6))
    resumen_mes.plot(kind='bar', stacked=True)
    plt.title("Cambios por Tipo y Mes")
    plt.ylabel("Cantidad de Cambios")
    plt.xlabel("Mes")
    plt.xticks(rotation=45)
    plt.tight_layout()

    temp_dir = tempfile.gettempdir()
    grafico_path = os.path.join(temp_dir, "cambios_por_mes.png")
    plt.savefig(grafico_path)
    plt.close()

    top_columnas = historial_df["Columna"].value_counts().head(10)
    plt.figure(figsize=(10, 6))
    sns.barplot(x=top_columnas.values, y=top_columnas.index, palette="viridis")
    plt.title("Top 10 Columnas con Más Cambios")
    plt.xlabel("Cantidad de Cambios")
    plt.ylabel("Columna")
    plt.tight_layout()

    grafico2_path = os.path.join(temp_dir, "top_columnas_cambios.png")
    plt.savefig(grafico2_path)
    plt.close()

    wb = load_workbook(archivo_excel)
    if "Gráficos" in wb.sheetnames:
        ws = wb["Gráficos"]
    else:
        ws = wb.create_sheet("Gráficos")

    img1 = Image(grafico_path)
    img2 = Image(grafico2_path)

    ws.add_image(img1, "A1")
    ws.add_image(img2, "A20")

    wb.save(archivo_excel)

    os.remove(grafico_path)
    os.remove(grafico2_path)

    log_func and log_func("Gráficos generados y añadidos al archivo Excel.")
